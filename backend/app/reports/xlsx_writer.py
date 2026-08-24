"""Workbook rendering, via the optional ``openpyxl`` dependency."""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any

from app.core.exceptions import ServiceUnavailableError
from app.reports.table import Document, Table, to_native, to_text

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    AVAILABLE = True
except ImportError:  # pragma: no cover - depends on what is installed
    AVAILABLE = False

HEADER_COLOUR = "6D28D9"
MAX_SHEET_NAME = 31
# Sampled rather than scanned in full: column widths from the first couple of
# hundred rows are indistinguishable from widths taken over fifty thousand,
# and the full scan is the slowest part of a large export.
WIDTH_SAMPLE = 200


def render(document: Document) -> bytes:
    if not AVAILABLE:  # pragma: no cover - depends on what is installed
        raise ServiceUnavailableError(
            "Excel export is not available on this server. Export as CSV instead."
        )

    workbook = Workbook()
    workbook.remove(workbook.active)

    for table in document.tables:
        _add_table(workbook, table, document.timezone)
    _add_cover(workbook, document)

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _add_table(workbook: Any, table: Table, timezone: str) -> None:
    sheet = workbook.create_sheet(_sheet_name(table.name, workbook.sheetnames))
    sheet.append([column.label for column in table.columns])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOUR)
        cell.alignment = Alignment(horizontal="center")

    for row in table.rows:
        # Native types, not strings: a score written as text cannot be sorted,
        # averaged or charted, which is most of why a teacher asked for a
        # workbook instead of a CSV.
        sheet.append([to_native(value, timezone) for value in table.values(row)])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for index, column in enumerate(table.columns, start=1):
        letter = get_column_letter(index)
        widest = max(
            [len(column.label)]
            + [len(to_text(row.get(column.key), timezone)) for row in table.rows[:WIDTH_SAMPLE]]
        )
        sheet.column_dimensions[letter].width = min(48, max(10, widest + 2))

        # Chosen from the source values, not from the cells. openpyxl widens a
        # `date` to a `datetime` on the way in, so by the time a cell is read
        # back a trend date is indistinguishable from a timestamp — and would
        # be formatted with a 00:00 on every row.
        fmt = _date_format(table, column.key)
        if fmt:
            for cell in sheet[letter][1:]:
                cell.number_format = fmt

    if table.note:
        sheet.append([])
        sheet.append([table.note])


def _date_format(table: Table, key: str) -> str | None:
    """The number format for a column, or None if it holds no dates."""
    for row in table.rows[:WIDTH_SAMPLE]:
        value = row.get(key)
        if isinstance(value, datetime):
            return "yyyy-mm-dd hh:mm"
        if isinstance(value, date):
            return "yyyy-mm-dd"
    return None


def _add_cover(workbook: Any, document: Document) -> None:
    """A first sheet saying what this file is.

    Inserted at the front because the reader needs the scope and the period
    before the numbers; a workbook of bare tables invites exactly the
    misreading the header prevents.
    """
    cover = workbook.create_sheet("About", 0)
    cover["A1"] = document.title
    cover["A1"].font = Font(bold=True, size=14)

    row = 2
    if document.subtitle:
        cover[f"A{row}"] = document.subtitle
        row += 1
    generated = to_text(document.generated_at, document.timezone)
    for label, value in ({"Generated": generated} | document.meta).items():
        cover[f"A{row}"] = label
        cover[f"A{row}"].font = Font(bold=True)
        cover[f"B{row}"] = value
        row += 1

    cover.column_dimensions["A"].width = 24
    cover.column_dimensions["B"].width = 56


def _sheet_name(name: str, taken: list[str]) -> str:
    """Excel rejects sheet names over 31 characters or containing ``[]:*?/\\``."""
    cleaned = "".join("-" if character in r"[]:*?/\\" else character for character in name)
    cleaned = cleaned[:MAX_SHEET_NAME] or "Sheet"
    if cleaned not in taken:
        return cleaned

    for suffix in range(2, 100):
        candidate = f"{cleaned[: MAX_SHEET_NAME - len(str(suffix)) - 1]} {suffix}"
        if candidate not in taken:
            return candidate
    raise ValueError("Too many sheets sharing one name.")  # pragma: no cover
